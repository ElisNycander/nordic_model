"""
Compute various parameters needed for model, such as production statistics

"""

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from contextlib import redirect_stdout

from model_definitions import country_to_areas, area_to_country, all_areas, cm_per_inch, bidz2maf_pecd
import entsoe_transparency_db
from model_definitions import entsoe_type_map
from pathlib import Path
from maf_hydro_data import get_maf_weights
import datetime

data_path = 'D:/NordicModel/InputData'

def compute_average_price(areas=['SE1','DK1'],startyear=2016,endyear=2017):

    starttime = f'{startyear}0101:00'
    endtime = f'{endyear}1231:23'

    db = entsoe_transparency_db.Database(db='D:/Data/entsoe_prices.db')
    prices = db.select_price_data(areas=areas,starttime=starttime,endtime=endtime)
    avg_price = {}
    for area in prices.columns:
        avg_price[area] = np.mean(prices[area])
    return avg_price

def compute_load_shares(year=2016,countries=['SE','DK','NO'],data_path='D:/NordicModel/InputData',db_path='D:/Data'):
    """ Compute load shares by country over entire year,
    Usage:
        allocate aggregate solar capacity by area
    """
    areas = []
    for c in countries:
        for a in country_to_areas[c]:
            areas.append(a)

    load_shares = pd.Series(index=[a for a in all_areas if area_to_country[a] in countries],dtype=float)
    db = entsoe_transparency_db.Database(db=Path(db_path)/'load.db')
    consumption = db.select_load_data(areas=areas,starttime=f'{year}0101:00',endtime=f'{year}1231:23')

    for c in countries:
        tot_load = consumption.loc[:,country_to_areas[c]].sum().sum()
        for a in country_to_areas[c]:
            load_shares.at[a] = consumption.loc[:,a].sum()/tot_load

    load_shares.to_excel(Path(data_path) / f'load_shares.xlsx')
    return load_shares

def compute_wind_generation(areas=['SE','DK'],year=2018,aggregate_countries=True):
    """ Compute aggregate wind generation in TWh for all countries the given year """
    if aggregate_countries:
        countries = areas
        areas = []
        for c in countries:
            for a in country_to_areas[c]:
                areas.append(a)
    else:
        countries = areas

    # get yearly wind generation
    db = entsoe_transparency_db.Database(db='D:/Data/entsoe_gen.db')
    data = db.select_gen_per_type_wrap_v2(starttime=f'{year}0101:00',endtime=f'{year}1231:23',areas=areas,type_map={'Wind':['Wind offsh','Wind onsh']})
    # wind2 = db2.select_gen_per_type_data(areas=self.areas,types=['Wind onsh','Wind offsh'],starttime='{0}0101:00'.format(year),endtime='{0}1231:23'.format(year))

    wind_yearly = pd.Series(index=countries)
    for c in wind_yearly.index:
        if aggregate_countries:
            wind_yearly.at[c] = sum(data[a]['Wind'].sum() for a in country_to_areas[c])/1e6
        else:
            wind_yearly.at[c] = data[a]['Wind'].sum() / 1e6
    return wind_yearly


def get_entsoe_production_stats(startyear=2016,endyear=2016,areas=['DE','FI'],limit=50,filename='gen_stats.xlsx',
                                data_path='D:/NordicModel/Data',db_path='D:/Data'):
    """ Use entsoe database to get minimum and maximum production
    values for different production types
    """
    path = Path(data_path)
    path.mkdir(exist_ok=True,parents=True)


    db = entsoe_transparency_db.Database(db=Path(db_path)/'gen.db')
    with pd.ExcelWriter(Path(data_path) / filename) as writer:
        for year in range(startyear,endyear+1):
            starttime = f'{year}0101:00'
            endtime = f'{year}1231:23'

            data = db.select_gen_per_type_wrap_v2(starttime=starttime,endtime=endtime,areas=areas,type_map=entsoe_type_map,drop_data=True,limit=limit)

            stats = pd.DataFrame(dtype=float,index=['min','max','avg','maxramp','minramp'],
                 columns=pd.MultiIndex.from_product([areas,[g for g in entsoe_type_map]],names=['area','type']))

            for area,gtype in stats.columns:
                # print(f'{area},{gtype}')
                if gtype in data[area]:
                    stats.at['min',(area,gtype)] = np.min(data[area][gtype])
                    stats.at['max',(area,gtype)] = np.max(data[area][gtype])
                    stats.at['avg',(area,gtype)] = np.mean(data[area][gtype])

                    ramp = data[area][gtype].diff()
                    stats.at['minramp',(area,gtype)] = np.min(ramp)
                    stats.at['maxramp',(area,gtype)] = np.max(ramp)

            stats.to_excel(writer,sheet_name=f'{year}')
    return stats,data

def capacity_vs_maxprod(year = 2019,redo_stats = False,show_title = True,eps_fig = False):
    """
    Compare the ENTSO-E capacity with maximum production values, for generation per type
    :return:
    """
    # year = 2019
    # redo_stats = False

    if redo_stats:
        stats = get_entsoe_production_stats(startyear=year,endyear=year,areas=all_areas,limit=50)
    else:
        stats = pd.read_excel(Path(data_path)/f'gen_stats.xlsx',index_col=0,header=[0,1])
    cap = get_entsoe_capacity(areas=all_areas,year=year)

    #%%
    large_areas = ['GB','PL','DE','NL']
    # show_title = True
    # eps_fig = False
    fig_path = Path(data_path) / 'Figures'
    fig_path.mkdir(exist_ok=True,parents=True)
    """
    Compare ENTSO-E capacity values with maximum production stats
    
    Print latex tables and figures with capacity and generator info
    """
    fig_size = (16/cm_per_inch,8/cm_per_inch)
    areas = all_areas
    # summarize thermal production
    thermal_data = pd.DataFrame(index=areas,columns=['pmax','capacity','diff'])
    for area in areas:
        thermal_data.at[area,'capacity'] = cap.at[area,'Thermal']
        thermal_data.at[area,'pmax'] = stats.at['max',(area,'Thermal')]
        thermal_data.at[area,'diff'] = thermal_data.at[area,'capacity'] - thermal_data.at[area,'pmax']
    thermal_data = thermal_data.fillna(0)

    # summarize hydro production
    hydro_data = pd.DataFrame(index=areas,columns=['pmax','capacity','diff'])
    for area in thermal_data.index:
        hydro_data.at[area,'capacity'] = cap.at[area,'Hydro']
        hydro_data.at[area,'pmax'] = stats.at['max',(area,'Hydro')]
        hydro_data.loc[area,'diff'] = hydro_data.at[area,'capacity'] - hydro_data.at[area,'pmax']
    hydro_data = hydro_data.fillna(0)

    f = plt.figure()
    ax = f.add_subplot(1,1,1)
    areas1 = [a for a in areas if a not in large_areas]
    areas2 = [a for a in areas if a in large_areas]

    for i,plot_areas in enumerate([areas1,areas2]):
        ax.cla()
        thermal_data.loc[plot_areas,['pmax','capacity']].plot.bar(ax=ax)
        plt.grid()
        if show_title:
            plt.title('Thermal capacity')
        plt.ylabel('MW')
        plt.gcf().set_size_inches(fig_size)
        plt.tight_layout()
        plt.savefig(fig_path/f'thermal_capacity_{i}.png')
        if eps_fig:
            plt.savefig(fig_path/f'thermal_capacity_{i}.eps')


        ax.cla()
        hydro_data.loc[plot_areas,['pmax','capacity']].plot.bar(ax=ax)
        plt.grid()
        if show_title:
            plt.title('Hydro capacity')
        plt.ylabel('MW')
        plt.gcf().set_size_inches(fig_size)
        plt.tight_layout()
        plt.savefig(fig_path/f'hydro_capacity_{i}.png')
        if eps_fig:
            plt.savefig(fig_path/f'hydro_capacity_{i}.eps')

def get_transmission_capacity():
    """ Get maximum values for transmission capacity """

    db = entsoe_transparency_db.Database(db='D:/Data/entsoe_exchanges.db')
    area_sep = '->'
    conn = []
    for i,a in enumerate(all_areas):
        for aa in all_areas[i+1:]:
            conn.append(f'{a}{area_sep}{aa}')
            conn.append(f'{aa}{area_sep}{a}')

    df = db.select_flow_data(starttime='20180101:00',endtime='20201231:00',table='capacity',connections=conn)
    nom_cap = df.max()
    nom_cap.to_excel(Path(data_path) / 'capacity.xlsx')

def calculate_hydro_production():
    """ Calculate yearly hydro """
    db = entsoe_transparency_db.Database(db='D:/Data/entsoe_gen.db')

    countries = ['PL','DE','GB','NL']
    gen = db.select_gen_per_type_wrap_v2(starttime='20180101:00',endtime='20191231:23',areas=countries,drop_data=False)

    #%%
    hydro = pd.DataFrame(dtype=float,index=countries,columns=['pump','nopump'])
    for c in countries:
        hydro.at[c,'pump'] = gen[c]['Hydro'].sum()
        hydro.at[c,'nopump'] = gen[c].loc[:,['Hydro ror','Hydro res']].sum().sum()
    hydro = hydro / 1e3 / 52 / 2
    print(hydro)
    return hydro


def get_maf_pump_capacity():
    """ Get pumping capacity from maf data """

    from maf_hydro_data import Database
    maf_db = Database(db='D:/Data/entsoe_maf_hydro.db')
    bidz2maf_w, maf_areas = get_maf_weights(areas=all_areas)
    maf_cap = maf_db.get_capacity()

    pump = pd.Series(0,dtype=float,index=all_areas)

    for a in pump.index:
        for ma,w in bidz2maf_w[a]:
            if ma in maf_cap.index:
                pump.at[a] += w*maf_cap.loc[ma,['POLPUMP','PCLPUMP']].sum()
    return pump

def fit_reservoir_capacity(effective_range=False):
    """
    Use reservoir capacity from reservoir database (in Million cubic meter) and given values (in TWh) from Nordpool
    to get estimates of the total reservoir capacity for remaining countries
    :return:
    """
    #%% compare nordpool reservoir capacities and reservoir capacities from reservoir database
    from power_plants import Database as PlantDatabase

    db = PlantDatabase(db='D:/Data/power_plants.db')


    df = db.select_data(table='reservoirs',select_column='country',
                        column_vals=['United Kingdom','Germany','Poland','Sweden','Norway','Finland','Latvia','Lithuania'])
    df = df.loc[df.Hydroelectricity == 'x',:]

    df.drop(columns=[c for c in df.columns if c not in ['Country','Name_of_dam','Reservoir_capacity']],inplace=True)

    cap = df.groupby(['Country']).sum().loc[:,'Reservoir_capacity']
    cap.index = ['FI','DE','LV','LT','NO','PL','SE','GB']

    if effective_range:
        reservoir_capacity = { # GWh
            'SE1':11326,
            'SE2':13533,
            'SE3':1790,
            'SE4':180,
            'FI':2952,
            'NO1':6078,
            'NO2':21671,
            'NO3':7719,
            'NO4':14676,
            'NO5':14090,
            'LT':11.8,
            'LV':9.4,
        }
    else:
        reservoir_capacity = {
            'NO1':5787,
            'NO2':32725,
            'NO3':7809,
            'NO4':19367,
            'NO5':16523,
            'SE1':14810,
            'SE2':15730,
            'SE3':2911,
            'SE4':224,
            'FI':5530,
            'LT':12.2,
            'LV':11.2,
        }

    from model_definitions import country_to_areas

    df_cap = pd.DataFrame(dtype=float,index=['SE','FI','NO','LV','LT'],columns=['nordpool','reservoir'])
    for c in df_cap.index:
        df_cap.at[c,'nordpool'] = sum(reservoir_capacity[a] for a in country_to_areas[c])
        df_cap.at[c,'reservoir'] = cap.at[c]

    p = np.polyfit(x=df_cap.reservoir,y=df_cap.nordpool,deg=1)

    xvals = cap.loc[[c for c in cap.index if c not in df_cap.index]]
    yvals = np.polyval(p,xvals)

    #%%
    f,ax = plt.subplots()
    plt.plot(df_cap.reservoir,df_cap.nordpool,'*',label='Known')
    plt.plot(xvals,yvals,'o',label='Unknown')
    xx = [min(cap),max(cap)]
    plt.plot(xx,np.polyval(p,xx),'k--',label='fit')
    for x,y,name in zip(np.array(xvals),yvals,xvals.index):
        plt.text(x,y,name)
    for x,y,name in zip(df_cap.reservoir,df_cap.nordpool,df_cap.index):
        plt.text(x,y,name)
    plt.ylabel('Nordpool reservoir capacity (GWh)')
    plt.xlabel('Aquastat reservoir capacity (Mm3)')
    plt.legend()
    plt.grid()

    res = pd.Series(data=yvals,index=xvals.index)
    res.name = 'GWh'
    if effective_range:
        res.to_excel(Path(data_path) / f'reservoir_capacity_effective.xlsx')
    else:
        res.to_excel(Path(data_path) / f'reservoir_capacity.xlsx')
    return res

def get_effective_reservoir_range():
    db = entsoe_transparency_db.Database(db='D:/Data/entsoe_reservoir.db')
    #
    areas = ['SE1','SE2','SE3','SE4','FI','NO1','NO2','NO3','NO4','NO5','LT','LV']

    res_max = db.select_max(table_type='reservoir',areas=areas)
    res_min = db.select_max(table_type='reservoir',areas=areas,get_min=True)

    df = pd.DataFrame(index=areas,columns=['max','min'],dtype=float)
    df['max'] = res_max
    df['min'] = res_min
    df['eff'] = df['max']-df['min']

    df.to_excel(Path(data_path) / f'effective_reservoir_range.xlsx')
    return df

def reservoir_capacity_maf_vs_entsoe_check():
    #%% compare nordpool reservoir capacities and reservoir capacities from reservoir database
    from power_plants import Database as PlantDatabase
    from entsoe_transparency_db import Database as EntsoeDatabase
    from maf_hydro_data import Database as MafDatabase

    country2iso = {
        'NO':'NOR',
        'SE':'SWE',
        'FI':'FIN',
        'DK':'DNK',
        'PL':'POL',
        'DE':'DEU',
        'GB':'GBR',
        'NL':'NLD',
        'EE':'EST',
        'LV':'LVA',
        'LT':'LTU'
    }
    iso2country = {}
    for c in country2iso:
        iso2country[country2iso[c]] = c

    # countries for plant database
    countries = []
    for a in all_areas:
        if area_to_country[a] not in countries:
            countries.append(area_to_country[a])


    plant_db = PlantDatabase(db='D:/Data/power_plants.db')
    df = plant_db.select_data(table='reservoirs',select_column='ISO_alpha_3',
                              column_vals=[country2iso[c] for c in countries])
    df = df.loc[df.Hydroelectricity == 'x',:]

    df.drop(columns=[c for c in df.columns if c not in ['Name_of_dam','ISO_alpha_3','Reservoir_capacity',
                                                        'Decimal_degree_latitude','Decimal_degree_longitude']],inplace=True)

    df.columns = ['name','country','capacity','lat','lon']

    #%% map dams to price areas
    df['area'] = df['country']

    rcap = df.groupby(by=['country']).sum().loc[:,['capacity']]
    rcap.index = [iso2country[c] for c in rcap.index]



    #%% get min/max reservoir levels
    entsoe_db = EntsoeDatabase(db='D:/Data/entsoe_reservoir.db')
    rmax = entsoe_db.select_max(table_type='reservoir',areas=all_areas)
    rmin = entsoe_db.select_max(table_type='reservoir',areas=all_areas,get_min=True)
    rmax = rmax[ rmax > 0]
    for c in ['SE','NO']:
        rmax.at[c] = rmax.loc[country_to_areas[c]].sum()
        rmin.at[c] = rmin.loc[country_to_areas[c]].sum()
    #%% get maf capacity
    maf_db = MafDatabase(db='D:/Data/entsoe_maf_hydro.db')
    mafcap = maf_db.get_capacity()
    bidz2maf,maf_areas = get_maf_weights(areas=all_areas)
    mafcap = mafcap.loc[[ma for ma in maf_areas if ma in mafcap.index],:]
    mafcap['tot'] = mafcap.loc[:,['RES','ROR','POL','PCL']].sum(axis=1)
    # for c in ['SE','NO']:

    # map maf areas capacity to bid zone capacities
    mafcap2 = pd.Series(0.0,dtype=float,index=all_areas)
    for a in mafcap2.index:
        if a in bidz2maf:
            for ma,w in bidz2maf[a]:
                if ma in mafcap.index:
                    mafcap2.at[a] += w*mafcap.at[ma,'tot']


    #%% plot maf capacity vs maximum reservoir content
    eff_range = False
    show_text = True
    plot_areas = ['SE1','SE2','SE3','SE4','NO1','NO2','NO3','NO4','NO5','FI','LV','LT']

    if eff_range:
        x = np.array(rmax.loc[plot_areas]-rmin.loc[plot_areas])
    else:
        x = np.array(rmax.loc[plot_areas])

    y = np.array(mafcap2.loc[plot_areas])

    f,ax = plt.subplots()
    ax.scatter(x,y,label='data')
    plt.xlabel('Maximum reservoir level (GWh)')
    plt.ylabel('MAF reservoir capacity (GWh)')
    ax.grid()
    if show_text:
        for i,a in enumerate(plot_areas):
            plt.text(x[i],y[i],a)

    lfit = np.polyfit(x,y,deg=1)
    xvals = [np.min(x),np.max(x)]
    plt.plot(xvals,np.polyval(lfit,xvals),'r-',label='fit')
    plt.legend(title=f'slope: {lfit[0]:0.3f}')
    plt.savefig(f'D:/NordicModel/Figures/reservoir_capacity_maf_vs_entsoe.png')

    return mafcap2

def estimate_co2_intensity():
    pass

    # rough EU ETS CO2 price
    co2_price = {
        2016:6,
        2017:7,
        2018:15,
        2019:25,
        2020:25,
        2021:50,
    }

    #%% use fitted costs for 2016 and 2019 to estimate CO2 intensity of thermal generation
    import pickle
    filepath = 'D:/NordicModel/InputData/costfit'
    fit_years = [2016,2019]

    pdic = {}
    costfit = {}
    for fit_year in fit_years:
        pdic[fit_year] = {}
        with open(Path(filepath) / f'{fit_year}/{fit_year}_fit.pkl','rb') as f:
            dic = pickle.load(f)
            for a in [k for k in dic.keys() if k not in ['binsize','starttime']]:
                pdic[fit_year][a] = [dic[a]['Thermal']['k'],
                                     dic[a]['Thermal']['mavg']]
            costfit[fit_year] = dic


    #%%
    fig_path = 'D:/NordicModel/InputData/Figures'
    co2_intensity_calc = {}
    price_eval = {
        'DE':40e3,
        'NL':8e3,
        'PL':15e3,
        'GB':25e3,
    }
    areas = ['DE','NL','PL','GB']
    pmax = 50e3
    # peval = 40e3

    f,ax = plt.subplots()

    for area in areas:
        ax.cla()
        xeval = price_eval[area]
        xvals = np.arange(xeval*1.5)
        for year in fit_years:
            yvals = np.polyval(pdic[year][area],xvals)
            plt.plot(xvals,yvals,label=f'{year}')
        plt.plot([xeval,xeval],[min(yvals),max(yvals)],'k--',label='eval')
        plt.legend()
        plt.grid()
        plt.title(f'{area}')
        plt.show()
        plt.savefig(Path(fig_path) / f'co2_intensity_est_{area}.png')

        price_diff = np.polyval(pdic[2019][area],xeval) - np.polyval(pdic[2016][area],xeval)
        # calculate co2 intensity
        co2_intensity_calc[area] = price_diff / (co2_price[2019]-co2_price[2016])

def draw_model_map():

    #%% make map with model areas
    fig_path = 'D:/NordicModel/Figures'
    hvdc_col = 'C1'
    hvdc_width = 1.5
    bidz_width = 0.8
    country_width = 0.3
    # country_color = 'khaki'
    country_color = 'wheat'

    # shapefile
    sname = 'ne_10m_admin_0_countries_lakes'
    spath = 'C:/Users/elisn/Box Sync/Data/countries'
    sfile = Path(spath) / sname

    # map options
    show_text = False
    proj = 'laea'
    width = 2.3e6
    height = 2.5e6
    lat_0 = 61
    lon_0 = 9.5

    # price zone cuts
    se1 = ((1.40735e6,1.86707e6),(1.69793e6,1.72759e6))
    se2 = ((1.295e6,1.39827e6),(1.56233e6,1.28204e6))
    se3 = ((1.29887e6,848115),(1.5662e6,925602))
    # no_x = (1.10393e6,1.27007e6)
    no_x = (1.10393e6,1.3e6)
    # no1 = ((1.10509e6,1.15443e6),no_x)
    no1 = ((1.10435e6,1.12919e6),no_x)

    # no2 = ((1.01217e6,1.20843e6),(1.20201e6,1.0922e6))
    no2 = ((970000,1.18e6),(1.20201e6,1.0922e6))
    no3 = ((1.00651e6,1.3516e6),no_x)
    no4 = ((1.27788e6,1.43246e6),no_x)
    no5 = ((1.37428e6,1.61209e6),(1.25113e6,1.69685e6))
    dk1 = ((1.1594e6,596283),(1.18652e6,689267))

    # HVDC connections
    hvdc_conn = {
        'SE3-FI':((1.61023e6,1.23518e6),(1.78657e6,1.29366e6)),
        'SE4-LT':((1.54628e6,743619),(1.8752e6,723518)),
        'SE4-PL':((1.48323e6,727173),(1.62485e6,563624)),
        'SE4-DE':((1.37633e6,634891),(1.37725e6,525249)),
        'DK2-DE':((1.31877e6,574588),(1.33887e6,508803)),
        'SE3-DK1':((1.21004e6,850520),(1.29593e6,866966)),
        'NO2-DK1':((1.13512e6,819455),(1.07116e6,932751)),
        'NO-NL':((969743,415608),(980708,930924)),
        'NO-DE':((998981,930010),(1.08944e6,455809)),
        'FI-EE':((1.9937e6,1.25697e6),(2.017e6,1.18707e6)),
        'GB-NL':((557085,247833),(775116,264867)),
        'GB-NO':((449808,689891),(953064,1.07665e6)),
        'NL-DK':((974061,413992),(1.09594e6,632596)),
    }
    under_construction = ['NL-DK']

    # label positions
    areas = {
        'SE1':(1.60107e6,1.9523e6),
        'SE2':(1.48097e6,1.55324e6),
        'SE3':(1.43835e6,1.16968e6),
        'SE4':(1.44222e6,801623),
        'NO1':(1.20589e6,1.2278e6),
        'NO2':(1.05091e6,1.06895e6),
        'NO3':(1.14777e6,1.42215e6),
        'NO4':(1.34536e6,1.91743e6),
        'NO5':(985050,1.26654e6),
        'DK1':(1.13227e6,734765),
        'DK2':(1.26788e6,635027),
        'NL':(888192,294085),
        'PL':(1.77929e6,336703),
        'DE':(1.17814e6,228221),
        'EE':(2.06212e6,1.11544e6),
        'LV':(2.11248e6,925602),
        'LT':(2.05409e6,728011),
        'FI':(1.95751e6,1.50675e6),
        'GB':(400025,383195),
    }

    #%%
    from mpl_toolkits.basemap import Basemap
    from matplotlib.patches import Polygon
    from matplotlib.collections import PatchCollection

    #%% make plot
    m = Basemap(projection=proj,resolution='l',
                width=width, height=height,
                lat_0=lat_0, lon_0=lon_0)


    f,ax = plt.subplots()
    f.set_size_inches(6,6)



    #% fill countries
    m.readshapefile(sfile.__str__(), 'units', color='#444444', linewidth=country_width)

    iso_countries = ['SWE','NOR','POL','FIN','DNK','DEU','GBR','NLD','EST','LVA','LTU']
    for info, shape in zip(m.units_info, m.units):
        iso3 = info['ADM0_A3']
        if iso3 in iso_countries:
            patches = [Polygon(np.array(shape), True)]
            pc = PatchCollection(patches)
            pc.set_facecolor(country_color)
            ax.add_collection(pc)

    # plot cuts
    for cut in [se1,se2,se3,no1,no2,no3,no4,no5,dk1]:
        xvals = [cut[0][0],cut[1][0]]
        yvals = [cut[0][1],cut[1][1]]
        plt.plot(xvals,yvals,'k',linewidth=bidz_width)
    # plot hvdc connections
    leg1 = False
    leg2 = False
    for conn in hvdc_conn:
        if conn in under_construction:
            linestyle = '--'
            if not leg1:
                label = 'HVDC under constr'
                leg1 = True
            else:
                label = None
        else:
            linestyle = '-'
            if not leg2:
                label = 'HVDC'
                leg2 = True
            else:
                label = None
        c = hvdc_conn[conn]
        xvals = [c[0][0],c[1][0]]
        yvals = [c[0][1],c[1][1]]
        plt.plot(xvals,yvals,color=hvdc_col,linewidth=hvdc_width,linestyle=linestyle,label=label)

    # plot area labels
    for a in areas:
        plt.text(areas[a][0],areas[a][1],a,color='black',fontsize=10,ha='center',va='center',bbox=dict(
            fc='none',ec='none',pad=0.05))

    plt.legend(loc='upper left')
    plt.savefig(Path(fig_path)/f'model_map.png')
    plt.savefig(Path(fig_path)/f'model_map.eps')

def compare_maf_vs_entsoe_inflow_data(inflow_db='inflow2.db',db_path='D:/Data',inflow_table='inflow',fig_tag='f1'):

    pass
    #%% MAF and Nordpool inflow data
    from maf_hydro_data import Database as MafDB
    titles = True
    fig_path = 'D:/NordicModel/Figures'
    starttime = '20150107:00'
    endtime = '20161225:00'
    areas = ['SE1','SE2','SE3','SE4','NO1','NO2','NO3','NO4','NO5','FI']
    # from model_definitions import nordic_areas
    entsoe_db = entsoe_transparency_db.Database(db=Path(db_path)/inflow_db)

    inflow_entsoe = entsoe_db.select_inflow_data_v2(starttime=starttime,endtime=endtime,areas=areas,table=inflow_table,date_index=False)
    inflow_entsoe[inflow_entsoe < 0] = 0
    #%% maf inflow data

    sfmt_in = '%Y%m%d:%H'
    sfmt_out = '%Y%m%d'
    maf_db = MafDB('D:/Data/maf_hydro.db')

    inflow_maf = maf_db.get_weekly_inflow_bidz(starttime=starttime,endtime=endtime,areas=areas,map_path='D:/NordicModel/InputData')

    rmse = (inflow_maf - inflow_entsoe).abs().mean()
    rmse_n = rmse / inflow_entsoe.mean()

    f,ax = plt.subplots()
    for a in areas:
        ax.cla()
        inflow_entsoe[a].plot(ax=ax,label='entsoe')
        inflow_maf[a].plot(ax=ax,label='maf')
        plt.grid()
        ax.set_ylabel('Inflow (GWh)')
        ax.legend(title=f'Norm. MAE: {rmse_n[a]:0.4f}')
        if titles:
            ax.set_title(a)
        plt.savefig(Path(fig_path) / f'maf_inflow_validatate_{fig_tag}_{a}.png')
    pass


def inflow_energiforetagen():
    """
    Read inflow data from energiforetagen, considering iso calendar format
    :return:
    """
    from week_conversion import WeekDef
    wd = WeekDef(week_start=4,proper_week=True) # this data uses iso weeks

    from help_functions import week2date_iso, week2str
    import openpyxl
    file_path = 'C:/Users/elisn/Box Sync/Nordic490/Data/Inflow_SE_2000-2018.xlsx'

    wb = openpyxl.load_workbook(file_path)
    ws = wb['Inflow - Sweden']

    cidxs = range(3, 22)
    w1 = '2000:01'
    w2 = '2018:52'
    index = wd.range2weeks(w1,w2)

    inflow_SE = pd.DataFrame(dtype=float, index=index, columns=['SE'])
    hrow = 8  # row with headers
    for cidx in cidxs:
        year = ws.cell(hrow, cidx).value
        for ridx in range(hrow + 1, hrow + 54):
            week = str(ws.cell(ridx, 2).value).zfill(2)
            val = ws.cell(ridx, cidx).value
            if not val is None:
                inflow_SE.at[f'{year}:{week}','SE'] = val
    return inflow_SE



if __name__ == "__main__":


    draw_model_map()
    def main():
        pass
    # compare_maf_vs_entsoe_inflow_data('inflow2.db',fig_tag='inflow2')





