"""
Create database with hydro model data from ENTSO-E MAF 20202
Note: For weekly values, MAF counts the number of weeks as the number of Sundays in a year
"""

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import datetime
import sqlite3
from pathlib import Path
from help_functions import str_to_date
from help_functions import weekdd
from week_conversion import WeekDef
import os

# maf counts weeks as starting with Monday
weekDef = WeekDef(week_start=7,proper_week=False)



maf_data_path = 'D:/Data/MAF/Hydro data'

def _time2table_(timestr,table_type='flow'):
    """ Given timestr of format 'YYYYMMDD:HHMM' produce string of format 'YYYYMM'
    This function is used to map a time to a table, to be able to divide data
    into multiple tables for faster search. May be updated if further data
    division is required, e.g. one table per day
    """
    if table_type == 'flow' or table_type == 'exchange' or table_type == 'capacity':
        return timestr[0:4]  # currently one table per year
    else: # return year by default
        return timestr[0:4]

def _create_table_(c, name,table_type='flow'):
    """ Drop table if it exists, then create new table with given name """

    c.execute(f'DROP TABLE IF EXISTS {name}')

    if table_type == 'flow':
        c.execute(f'CREATE TABLE {name} (' + \
                  'time TEXT NOT NULL,' + \
                  'value REAL NOT NULL' + \
                  ')')
    elif table_type == 'gen_per_type_v2':
        c.execute(f'CREATE TABLE {name} (' + \
                  'time TEXT NOT NULL,' + \
                  'type TEXT NOT NULL,' + \
                  'value REAL NOT NULL' + \
                  ')')
    else:
        c.execute(f'CREATE TABLE {name} (' + \
                  'time TEXT NOT NULL,' + \
                  'value REAL NOT NULL' + \
                  ')')
    # print(f'_create_table_(): Unknown table_type "{table_type}"')

def _execute_(c, cmd):
    try:
        c.execute(cmd)
    except sqlite3.Error as e:
        print('Executing command ''{0}'' returned error: {1}'.format(cmd, e))



class Database():

    def __init__(self,db='Data/maf_hydro.db',maf_path='D:/Data/MAF/Hydro data',write=False):

        if not write and not os.path.isfile(db):
            # Read only, but file doesn't exist
            raise FileNotFoundError(f"Database {db} not found")

        self.db = db
        self.maf_path = maf_path

        if Path(self.db).exists():
            pass

        self.hydro_types = {
            'ROR':'Run-of-River and pondage',
            'RES':'Reservoir',
            'POL':'Pump storage - Open Loop',
            'PCL':'Pump Storage - Closed Loop'
        }

    def get_distinct_values(self,table_name='capacity',field='area'):

        conn = sqlite3.connect(self.db)
        c = conn.cursor()

        cmd = f"SELECT DISTINCT {field} FROM {table_name}"
        _execute_(c,cmd)
        vals = []
        for r in c.fetchall():
            vals.append(r[0])

        conn.close()
        return vals

    def get_capacity(self):
        """ Select data from capacity table """
        #%% get capacities
        table_name = 'capacity'
        areas = self.get_distinct_values(table_name,'area')
        cols = self.get_distinct_values(table_name,'type')

        cap_df = pd.DataFrame(dtype=float,index=areas,columns=cols)

        conn = sqlite3.connect(self.db)
        c = conn.cursor()

        cmd = f"SELECT * FROM {table_name}"
        _execute_(c,cmd)
        for r in c.fetchall():
            cap_df.at[r[0],r[1]] = r[2]

        return cap_df

    def get_hydrotypes(self):
        """ Select data from hydrotypes table """
        #%% get capacities
        table_name = 'hydrotypes'

        areas = self.get_distinct_values(table_name,'area')
        cols = self.hydro_types
        # cols = self.get_distinct_values(table_name,'type')

        df = pd.DataFrame(False,dtype=bool,index=areas,columns=cols)

        conn = sqlite3.connect(self.db)
        c = conn.cursor()

        cmd = f"SELECT * FROM {table_name}"
        _execute_(c,cmd)
        for r in c.fetchall():
            df.at[r[0],r[1]] = True

        return df

    def create_gen_capacity_table(self):
        """ Create table with generation capacity data, from MAF_2020 excel file """

        conn = sqlite3.connect(self.db)
        cursor = conn.cursor()

        # create table
        table_name = f'gen_capacity'
        _execute_(cursor,f'DROP TABLE IF EXISTS {table_name}')
        _execute_(cursor,f'CREATE TABLE {table_name} (' + \
                  'area TEXT NOT NULL,' + \
                  'type TEXT NOT NULL,' + \
                  'year INT NOT NULL,' + \
                  'value REAL NOT NULL)' )

        conn.commit()


        for year in [2025,2030]:

            df = read_maf_excel_data(year=year)

            for c in df.columns:
                for itype in df.index:
                    val = df.at[itype,c]
                    if val > 0:
                        cmd = f'INSERT INTO {table_name} (area,type,year,value) VALUES ("{c}","{itype}",{year},{val})'
                        _execute_(cursor,cmd)
            conn.commit()
        conn.close()

    def create_capacity_table(self):
        """
        Create table with reservoir capacities, also check which types of inflow data exist and create
        table with different inflow types

        categories:
        run of river and pondage - ROR
        reservoir - RES
        pumped hydro open loop reservoir capacity - OL
        pumped hydro closed loop reservoir capacity - CL
        pumped hydro open loop pump capacity - OLPMP
        pumped hydro close loop pump capacity - CLPMP
        """
        import xlrd

        data_path = Path(self.maf_path)

        files = [f for f in os.listdir(data_path) if os.path.isfile(data_path/f) and 'xlsx' in f and '~' not in f]
        maf_areas =  [f.split('_')[1] for f in files]

        gen_df = pd.DataFrame(index=maf_areas,columns=[h for h in self.hydro_types]+['POLPUMP','PCLPUMP'])
        htype_df = pd.DataFrame(dtype=bool,index=maf_areas,columns=[h for h in self.hydro_types])

        for area,file in zip(maf_areas,files):
            print(area)
            wb = xlrd.open_workbook(data_path / file)
            # check which hydro data exists
            for htype in self.hydro_types:
                hname = self.hydro_types[htype]
                sheet = wb.sheet_by_name(hname)

                rrange = range(2,10)
                crange = range(1,10)
                has_data = False
                for ridx in rrange:
                    for cidx in crange:
                        if type(sheet.cell(ridx,cidx).value) is str:
                            break
                        elif sheet.cell(ridx,cidx).value > 0:
                            has_data = True
                            break
                htype_df.at[area,htype] = has_data

            # read capacity data
            sheet = wb.sheet_by_name('General Info')

            gen_df.at[area,'ROR'] = sheet.cell(2,1).value
            gen_df.at[area,'RES'] = sheet.cell(3,1).value
            gen_df.at[area,'POL'] = sheet.cell(4,1).value
            gen_df.at[area,'PCL'] = sheet.cell(5,1).value
            gen_df.at[area,'POLPUMP'] = sheet.cell(4,2).value
            gen_df.at[area,'PCLPUMP'] = sheet.cell(5,2).value

        #%% read data into database

        table_name = f'hydrotypes'

        conn = sqlite3.connect(self.db)
        c = conn.cursor()

        _execute_(c,f'DROP TABLE IF EXISTS {table_name}')
        _execute_(c,f'CREATE TABLE {table_name} (' + \
                  'area TEXT NOT NULL,' + \
                  'type TEXT NOT NULL)' )

        for area in htype_df.index:
            for htype in htype_df.columns:
                if htype_df.at[area,htype]:
                    _execute_(c,f'INSERT INTO {table_name} (area,type) VALUES ("{area}","{htype}")')


        conn.commit()

        #%% capacity data

        # create table with capacity data
        table_name = 'capacity'
        _execute_(c,f'DROP TABLE IF EXISTS {table_name}')
        _execute_(c,f'CREATE TABLE {table_name} (' + \
                  'area TEXT NOT NULL,' + \
                  'type TEXT NOT NULL,' + \
                  'value REAL NOT NULL)' )

        for area in gen_df.index:
            for col in gen_df.columns:
                if gen_df.at[area,col] > 0:
                    cmd = f'INSERT INTO {table_name} (area,type,value) VALUES ("{area}","{col}",{gen_df.at[area,col]})'
                    _execute_(c,cmd)

        conn.commit()
        conn.close()


    def create_tables(self,max_idx=None):
        """ 
        Put data from excel files into tables
        Create tables with inflow data, reservoir limits, generation limits etc 
        """
        # setup connection
        conn = sqlite3.connect(self.db)
        c = conn.cursor()
    
        #%%
        # for all hydro types, get first data field as inflow
        data_field_number = {
            'inflow':0,
            'mingen':5,
            'maxgen':6,
            'reslevel':9,
            'minres':10,
            'maxres':11,
        }
        hydrotypes_data = {
            'ROR':['inflow'],
            'RES':['inflow','mingen','maxgen','reslevel','minres','maxres'],
            'POL':['inflow','mingen','maxgen','reslevel','minres','maxres'],
            'PCL':['inflow','mingen','maxgen','reslevel','minres','maxres'],
        }
    
        #%% get capacities
        def_df = self.get_hydrotypes()
    
        data_path = Path(self.maf_path)
        files = [f for f in os.listdir(data_path) if os.path.isfile(data_path/f) and 'xlsx' in f and '~' not in f]
        maf_areas =  [f.split('_')[1] for f in files]
    
        #%%
        import xlrd

        if max_idx is not None:
            zip_areas = maf_areas[:max_idx]
            zip_files = files[:max_idx]
        else:
            zip_areas = maf_areas
            zip_files = files

        for area,file in zip(zip_areas,zip_files):
            # open excel file
            wb = xlrd.open_workbook(data_path / file)
    
            for htype in self.hydro_types:
                if area in def_df.index and def_df.at[area,htype]:
                    # open sheet
                    sheet = wb.sheet_by_name(self.hydro_types[htype])
    
                    tres = sheet.cell(1,0).value
                    if tres == 'Week':
                        time_resolution = 'w'
                    else:
                        time_resolution = 'd'
    
                    #%% use first row to find where new data block starts
                    data_block = []
                    ridx = 0
                    for cidx in range(sheet.ncols):
                        if type(sheet.cell(ridx,cidx).value) is str and sheet.cell(ridx,cidx).value.__len__() > 0:
                            data_block.append((cidx,sheet.cell(ridx,cidx).value))
    
                    #%%
    
                    # put data into table
                    for data_type in hydrotypes_data[htype]:
                        # data_type = 'inflow'
                        block_idx = data_field_number[data_type]
                        cstart = data_block[block_idx][0]
                        if block_idx < data_block.__len__() - 1:
                            cend = data_block[block_idx+1][0]
                        else:
                            cend = sheet.ncols
                        block_name = data_block[block_idx][1]
    
                        print(f'Reading for {area}: "{block_name}"')
    
                        # create table in database
                        table_name = f'{data_type}_{area}_{htype}'
                        has_values = False
    
                        cmd = f"DROP TABLE IF EXISTS {table_name}"
                        _execute_(c,cmd)
                        cmd = f'CREATE TABLE {table_name} (' + \
                              'time TEXT NOT NULL,' + \
                              'value REAL NOT NULL)'
                        _execute_(c,cmd)
                        conn.commit()
    
                        for cidx in range(cstart,cend):
                            year = int(sheet.cell(1,cidx).value)
                            date = datetime.datetime(year,1,1)
    
                            for ridx in range(2,sheet.nrows):
                                val = sheet.cell(ridx,cidx).value
                                if type(val) is float:
                                    if not has_values:
                                        has_values = True
                                    if time_resolution == 'd':
                                        sdate = date.strftime('%Y%m%d')
                                    else:
                                        sweek = f'{ridx-1}'.zfill(2)
                                        sdate = f'{year}:{sweek}'
                                    cmd = f"INSERT INTO {table_name} (time,value) VALUES ('{sdate}',{val})"
                                    _execute_(c,cmd)
                                if time_resolution == 'd':
                                    date += datetime.timedelta(days=1)
    
                        if not has_values: # drop empty tables (no data)
                            cmd = f"DROP TABLE IF EXISTS {table_name}"
                            _execute_(c,cmd)
                        conn.commit()

    def select_data(self,areas=[],htype='RES',data_type='inflow',starttime='20160101',endtime='20161010',wd=weekDef):
        """
        Select time series data

        Note: For the MAF hydro data, ENTSO-E counts the weeks of a year starting with Sunday as the first day,
        meaning that the years with 53 weeks are: [1984, 1989, 1995, 2000, 2006, 2012]
        For some countries (e.g. SE) all years have 53 weeks with values copied from week 52, while for some countries
        (e.g. PL) the years with 52 weeks have zero values for week 53
        """
        conn = sqlite3.connect(self.db)
        cursor = conn.cursor()

        if areas.__len__() == 0:
            get_areas = self.get_distinct_values('hydrotypes','area')
        else:
            get_areas = areas

        # starttime = '20160101'
        # endtime = '20161030'

        # get_areas = areas[:2]
        # area = areas[0]

        # htype = 'RES'
        # block = 'inflow'

        if data_type == 'inflow' and htype == 'ROR':
            time_res = 'd'
        else:
            time_res = 'w'

        #%% find relevant tables
        cmd = "SELECT name FROM sqlite_master WHERE type='table';"
        _execute_(cursor,cmd)

        rel_tables = [r[0] for r in cursor.fetchall() if data_type in r[0] and htype in r[0] and r[0].split('_')[1] in get_areas]

        if time_res == 'd':
            # daily resolution
            df = pd.DataFrame(index=pd.date_range(start=str_to_date(starttime),end=str_to_date(endtime),freq='D'),
                              columns=get_areas,
                              dtype=float)
            strstart = starttime
            strend = endtime
        else:
            ws = 7
            # weekly resolution
            index = wd.range2weeks(starttime,endtime,sout=True)
            df = pd.DataFrame(index=index,
                              columns=get_areas,
                              dtype=float)
            # strstart = date2week_v2(starttime,week_start=ws,str_out=True)
            # strend = date2week_v2(endtime,week_start=ws,str_out=True)
            strstart = index[0]
            strend = index[-1]

        #%%
        for table_name in rel_tables:
            area = table_name.split('_')[1]
            cmd = f"SELECT time,value FROM {table_name} WHERE time >= '{strstart}' and time <= '{strend}'"
            _execute_(cursor,cmd)
            for r in cursor.fetchall():
                if time_res == 'd':
                    df.at[str_to_date(r[0]),area] = r[1]
                else:
                    if r[0] in df.index:
                    # discard values not in index (i.e. additional weeks 53)
                        df.at[r[0],area] = r[1]
        return df

    def select_gen_capacity(self):
        """ Get whole table with generation capacities """

        conn = sqlite3.connect(self.db)
        cursor = conn.cursor()

        # find columns and rows
        table_name = 'gen_capacity'

        areas = self.get_distinct_values(table_name,'area')
        types = self.get_distinct_values(table_name,'type')
        types.sort()
        years = self.get_distinct_values(table_name,'year')

        df = pd.DataFrame(0.0,index=pd.MultiIndex.from_product([types,years],names=['type','year']),columns=areas)
        cmd = f'SELECT area,type,year,value from {table_name}'
        _execute_(cursor,cmd)
        for row in cursor.fetchall():
            df.at[(row[1],row[2]),row[0]] = row[3]

        gen_type_names = {
                # shorten some names
                'Hydro - Run of River & Pondage (Turbine Capacity)':'Hydro ror',
                'Hydro - Units with reservoir (Turbine Capacity)':'Hydro res',
                'Others non-renewable':'Other non-renew',
                'Others renewable':'Other renew',
                'Solar (Photovoltaic)':'Solar PV',
                'Solar (Thermal)':'Solar thermal',
        }
        stypes = []
        for t in types:
            if t in gen_type_names:
                stypes.append(gen_type_names[t])
            else:
                stypes.append(t)
        df.index = pd.MultiIndex.from_product([stypes,years])

        return df

    def select_inflow_bidz_wrap(self,starttime='20160101:00',endtime='20161230:23',
                               areas=('SE1','NO1','FI'),maf_path='D:/NordicModel/InputData',wd=weekDef,date_index=False):
        """ Get weekly inflow for given bid zones. Aggregates inflow values for Reservoir production and Open loop
        pumped hydro
        """
        if starttime.__len__() != endtime.__len__():
            raise ValueError("Starttime and endtime must be in same format")
        bidz2maf_w,maf_areas = get_maf_weights(areas,map_file = Path(maf_path) / 'maf_to_bidz_map.xlsx')

        # use RES and POL types for reservoir inflow
        inflow_maf_res = self.select_data(areas=maf_areas,htype='RES',data_type='inflow',starttime=starttime,endtime=endtime,wd=wd)
        inflow_maf_pol = self.select_data(areas=maf_areas,htype='POL',data_type='inflow',starttime=starttime,endtime=endtime,wd=wd)
        # drop nan columns
        inflow_maf_res.dropna(inplace=True,how='all',axis='columns')
        inflow_maf_pol.dropna(inplace=True,how='all',axis='columns')
        # also get ROR (daily data)
        if starttime.__len__() <= 7:
            # format is YYYY:MM, covert to day
            fmt = '%Y%m%d'
            sday = wd.week2date(starttime).strftime(fmt)
            eday = (wd.week2date(endtime) + datetime.timedelta(days=6)).strftime(fmt)
        else:
            sday = starttime
            eday = endtime
        inflow_maf_ror = self.select_data(areas=maf_areas,
                                        htype='ROR',data_type='inflow',
                                        starttime=sday,endtime=eday)
        inflow_maf_ror.dropna(inplace=True,how='all',axis='columns')

        inflow_maf = pd.DataFrame(0.0,columns=[a for a in bidz2maf_w],index=inflow_maf_res.index)
        inflow_daily = pd.DataFrame(0.0,columns=inflow_maf.columns,index=inflow_maf_ror.index)
        for a in inflow_maf.columns:
            for (ma,w) in bidz2maf_w[a]:
                if ma in inflow_maf_res.columns:
                    inflow_maf[a] += inflow_maf_res[ma]*w
                if ma in inflow_maf_pol.columns:
                    inflow_maf[a] += inflow_maf_pol[ma]*w
                if ma in inflow_maf_ror.columns:
                    inflow_daily[a] += inflow_maf_ror[ma]*w

        if date_index:
            # set index as date indicating start of week
            idx = pd.date_range(start=wd.week2date(inflow_maf.index[0]),
                                end=wd.week2date(inflow_maf.index[-1]),
                                freq='7D')
            if idx.__len__() != inflow_maf.shape[0]:
                raise ValueError('Incorrect length of date index! - indicates bug')
            inflow_maf.index = idx

        return inflow_maf, inflow_daily

def get_maf_weights(areas=['NO1','SE1'],map_file='D:/NordicModel/InputData/maf_to_bidz_map.xlsx'):
    """
    Get weights used for converting maf variables (inflow, capacities etc) to area quantities used in model
    For getting weights for areas which consist of several maf areas, the mapping in maf_to_bidz_map.xlsx
    is used. For areas not in this file, the mapping to a single maf area specified in bidz2maf_pecd is used
    """
    from model_definitions import bidz2maf_pecd
    maf2bidz_map_no = pd.read_excel(map_file,index_col=0)
    # map from bidz to (multiple) maf areas with corresponding weights

    bidz2maf_w = {}
    for a in maf2bidz_map_no.index:
        if a in areas:
            bidz2maf_w[a] = [(ma,maf2bidz_map_no.at[a,ma]) for ma in maf2bidz_map_no.columns if maf2bidz_map_no.at[a,ma] > 0]
    for a in bidz2maf_pecd:
        if a in areas and a not in bidz2maf_w:
            bidz2maf_w[a] = [(bidz2maf_pecd[a],1)]
    maf_areas = []
    for a in bidz2maf_w:
        for tup in bidz2maf_w[a]:
            ma = tup[0]
            if ma not in maf_areas:
                maf_areas.append(ma)
    return bidz2maf_w,maf_areas


def inflow_energiforetagen():
    #%% load inflow data from energiföretagen

    import openpyxl
    file_path = 'C:/Users/elisn/Box Sync/Nordic490/Data/Inflow_SE_2000-2018.xlsx'

    wb = openpyxl.load_workbook(file_path)
    ws = wb['Inflow - Sweden']

    years = range(2000, 2019)
    cidxs = range(3, 22)

    index = []
    for year in years:
        index = index + ['{0}:{1}'.format(year, week) for week in weekdd]

    inflow_SE = pd.DataFrame(dtype=float, index=index, columns=['SE'])
    hrow = 8  # row with headers
    for cidx in cidxs:
        year = ws.cell(hrow, cidx).value
        for ridx in range(hrow + 1, hrow + 54):
            week = ws.cell(ridx, 2).value
            val = ws.cell(ridx, cidx).value
            if not val is None:
                if week < 53:
                    # inflow_SE.loc[week,year] = val
                    inflow_SE.at['{0}:{1}'.format(year, weekdd[week - 1]), 'SE'] = val
                else:  # add week 53 to week 52
                    # inflow_SE.loc[52,year] = inflow_SE.loc[52,year] + val
                    inflow_SE.at['{0}:52'.format(year), 'SE'] = inflow_SE.at['{0}:52'.format(year), 'SE'] + val

    return inflow_SE

def maf_validate_data():

    db = Database(db='D:/Data/entsoe_maf_hydro.db')

    # db.create_tables()

    #%% select inflow data
    fig_path = 'C:/Users/elisn/Box Sync/Nordic490/Figures'
    # maf_areas = db.get_distinct_values('capacity','area')

    maf_area_map = {
        'SE':['SE01','SE02','SE03','SE04'],
        'NO':['NOS0','NOM1','NON1'],
        'FI':['FI00'],
    }

    maf_areas = ['SE01','SE02','SE03','SE04','FI00','NOS0','NOM1','NON1']
    starttime = '19820101'
    endtime = '20161231'

    ror = db.select_data(areas=maf_areas,starttime=starttime,endtime=endtime,htype='ROR')
    res = db.select_data(areas=maf_areas,starttime=starttime,endtime=endtime,htype='RES')
    pol = db.select_data(areas=maf_areas,starttime=starttime,endtime=endtime,htype='POL')

    # aggregate inflow data by countries, for comparison
    inflow_country = pd.DataFrame(dtype=float,index=res.index,columns=[c for c in maf_area_map])

    for c in ['SE','FI']:
        inflow_country[c] = res.loc[:,maf_area_map[c]].sum(axis=1)
    for c in ['NO']:
        inflow_country[c] = pol.loc[:,maf_area_map[c]].sum(axis=1)


    #%% compare MAF inflow with data from energiforetagen, for Sweden

    en_df = inflow_energiforetagen()

    from help_functions import week_to_date
    en_df.index = [week_to_date(i) for i in en_df.index]
    inflow_country.index = [week_to_date(i) for i in inflow_country.index]


    f,ax = plt.subplots()

    tidxs = [i for i in en_df.index if i in inflow_country.index]

    maf_plot_df = inflow_country.loc[tidxs,'SE']
    en_plot_df = en_df.loc[tidxs,'SE']

    maf_plot_df.plot(ax=ax,label='ENTSO-E MAF')
    en_plot_df.plot(ax=ax,label='Energiföretagen')

    rmse = np.mean(np.abs(en_plot_df-maf_plot_df))/np.mean(np.abs(en_plot_df))
    ax.legend(title=f'Normalized MAE: {rmse:0.4f}')
    ax.grid()
    ax.set_ylabel('GWh/week')
    plt.savefig(Path(fig_path) / 'maf_inflow_validate_energiforetagen.png')
    plt.savefig(Path(fig_path) / 'maf_inflow_validate_energiforetagen.eps')

    #%%


def read_maf_excel_data(year=2025,fpath = Path('D:/Data/MAF'),fname = 'MAF 2020 - Dataset.xlsx'):
    pass
    # year = 2025
    # fpath = Path('D:/Data/MAF')
    # fname = 'MAF 2020 - Dataset.xlsx'

    import xlrd
    wb = xlrd.open_workbook(fpath / fname)

    ws = wb.sheet_by_name(f'NT {year}')

    # scan columns
    areas = []
    for c in range(1,ws.ncols):
        areas.append(ws.cell(1,c).value)
    # scan rows
    types = []
    for r in range(2,ws.nrows):
        if ws.cell(r,0).value != '':
            types.append(ws.cell(r,0).value)
        else:
            break
    df = pd.DataFrame(dtype=float,index=types,columns=areas)
    for r in range(2,ws.nrows):
        if ws.cell(r,0).value == '':
            break
        rname = ws.cell(r,0).value
        for c in range(1,ws.ncols):
            cname = ws.cell(1,c).value
            df.at[rname,cname] = ws.cell(r,c).value

    return df

def map_NO_maf_areas(db_path='D:/Data',data_path='D:/NordicModel/Data'):
    """
    MAF uses 3 areas for NO, but Nordic model uses the 5 bidding zones. To use the MAF hydro data, the inflow must
    be distributed form the 3 MAF areas to the bidding zones. This is done using a mapping based on a database with
    hydro power plants, with known geographical location. The MAF area of each plant is determined based on its location,
    and since the bidding zone is contained in the plant data this creates a mapping based on the plant capacities.
    I.e., the inflow data for a maf area is distributed to the prize zones in proportion to the share of the total
    plant capacity of that area that is located within a bidding zone.

    Note: Relies on power plant database
    :return:
    """

    #%%
    from power_plants import Database as WriDB
    wri_db = WriDB(db=Path(db_path)/'power_plants.db')

    res_df = wri_db.select_data(table='reservoirs',select_column='country',column_vals=['Norway'])
    res_df.drop(columns=[c for c in res_df.columns if c not in ['Name_of_dam','Reservoir_capacity','Decimal_degree_latitude','Decimal_degree_longitude']],inplace=True)
    res_df.columns = ['name','cap','lat','lon']

    #%% load pypsa power plant data
    fpath = Path('C:/Users/elisn/Box Sync/Nordic490/Data/')
    fname = 'N490.xlsx'

    plant_df = pd.read_excel(fpath/fname,sheet_name='gen')
    plant_df = plant_df.loc[(plant_df.country=='NO') & (plant_df.type=='Hydro'),['name','bidz','Pmax','lat','lon']]

    res_df = plant_df
    res_df.columns = ['name','bidz','cap','lat','lon']

    #%% plot reservoirs
    show_text = False
    proj = 'laea'
    width = 1.3e6
    height = 1.6e6
    lat_0 = 64.5
    lon_0 = 16.5

    # borders between maf no areas, in x,y coordinates according to above projection
    cut1 = ((465380,730325),(333508,788725))
    cut2 = ((35854.4,466581),(425818,524981))

    from mpl_toolkits.basemap import Basemap
    map = Basemap(projection=proj,resolution='l',
                  width=width, height=height,
                  lat_0=lat_0, lon_0=lon_0)

    # plt.figure()
    f = plt.gcf()
    f.set_size_inches(4.5,6.5)
    f.clf()
    map.drawcountries()
    map.drawcoastlines()
    x,y=map(np.array(res_df.lon),np.array(res_df.lat))


    if show_text:
        for i in range(res_df.__len__()):
            plt.text(x[i],y[i],res_df.at[res_df.index[i],'name'])


    res_df['maf'] = 0
    res_df.loc[y<=765459,'maf'] = 1
    res_df.loc[y<=480643,'maf'] = 2
    # manual fixes
    res_df.loc[res_df.name=='Byafossen','maf'] = 0
    res_df.loc[res_df.name=='Hafsl.kr','maf'] = 2
    res_df.loc[res_df.name=='Vinstra','maf'] = 2
    res_df.loc[res_df.name=='Ovre vinstra','maf'] = 2

    a1_idx = res_df['maf'] == 0
    a2_idx = res_df['maf'] == 1
    a3_idx = res_df['maf'] == 2
    map.plot(x[a1_idx],y[a1_idx],color='b',marker='o',markersize=4,linestyle='None',label='NON1')
    map.plot(x[a2_idx],y[a2_idx],color='y',marker='o',markersize=4,linestyle='None',label='NOM1')
    map.plot(x[a3_idx],y[a3_idx],color='g',marker='o',markersize=4,linestyle='None',label='NOS0')
    plt.legend()
    map.plot([cut1[0][0],cut1[1][0]],[cut1[0][1],cut1[1][1]],color='k',linestyle='-',linewidth=2)
    map.plot([cut2[0][0],cut2[1][0]],[cut2[0][1],cut2[1][1]],color='k',linestyle='-',linewidth=2)
    plt.show()
    plt.savefig(Path(data_path)/'maf_bidz_map.png')

    #%% calculate weights
    wdf = pd.DataFrame(0.0,index=[f'NO{i}' for i in range(1,6)],columns=['NON1','NOM1','NOS0'])

    for maf_idx,maf_area in enumerate(wdf.columns):
        cap_tot = res_df.loc[res_df.maf==maf_idx,'cap'].sum()
        for bidz in wdf.index:
            wdf.at[bidz,maf_area] = res_df.loc[(res_df.maf==maf_idx)&(res_df.bidz==bidz),'cap'].sum() / cap_tot


    #%% validate capacity against entsoe-data
    from entsoe_transparency_db import Database as EntsoeDB
    entsoe_db = EntsoeDB(db='D:/Data/entsoe_cap_NO.db')
    entsoe_df = entsoe_db.select_cap_per_type_data(areas=wdf.index)
    hydro_types = ['Hydro Pumped Storage','Hydro Run-of-river and poundage','Hydro Water Reservoir']

    hydro_cap = pd.DataFrame(0.0,index=wdf.index,columns=['plants','entsoe'])
    for bidz in hydro_cap.index:
        hydro_cap.at[bidz,'plants'] = res_df.loc[res_df.bidz==bidz,'cap'].sum()
        hydro_cap.at[bidz,'entsoe'] = entsoe_df.loc[2020,(bidz,hydro_types)].sum()

    print('Compare total plant capacity to entsoe data:')
    print(hydro_cap)

    #%% save data
    wdf.to_excel(Path(data_path) / f'maf_to_bidz_map.xlsx')

if __name__ == "__main__":

    def main():
        pass

    pass


    #%%
    # inflow_se = inflow_energiforetagen_iso()



    db = Database(db='D:/Data/maf_hydro.db')


    #%%
    ws = 2
    areas=['PL00']
    htype='RES'
    data_type='inflow'
    starttime='20121220'
    endtime='20130110'

    # df = db.select_data(areas=['PL00'],htype='RES',starttime=starttime,endtime=endtime)

    df1,df2 = db.get_weekly_inflow_bidz(areas=['NO1','PL',],starttime='1982:01',endtime='2016:52')
    
    
